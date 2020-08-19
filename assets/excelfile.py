import openpyxl

wb = openpyxl.Workbook()
sheet = wb.active

col = []

def get_excel_sheet():
	wb = openpyxl.Workbook()
	sheet = wb.active

	# create variables
	sheet['A1'] = 'Serial No'
	sheet['B1'] = 'YouTube Video Link'
	sheet['C1'] = 'Video Views'
	sheet['D1'] = 'Uploaded Date'
	sheet['E1'] = 'Comments'
	sheet['F1'] = 'Likes'
	sheet['G1'] = 'Dislikes'

	wb.save(filename='test.xlsx')


def write_stats(file):
	wb = openpyxl.Workbook('test.xlsx')
	sheet = wb.active

	